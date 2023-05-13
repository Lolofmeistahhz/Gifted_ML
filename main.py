# region Import Region
import datetime
import numpy as np
import openpyxl
import timeit
import requests
import json
import time
import sys
import pandas as pd
from scipy.stats import rankdata

# endregion

# region Global Variables
students = []
communities = []

# endregion

# region Filename Settings
student_filename = "Data\\MainModel.xlsx"
communities_filename = "Data\\Communities.xlsx"

communities_result_filename = "Output\\CommunitiesResults.xlsx"
student_result_filename = "Output\\StudentsResults.xlsx"


# endregion

# region Classification Methods
def get_classification_values(arr):
    return {
        "max": np.percentile(arr, 75),
        "min": np.percentile(arr, 24)
    }


def classification(prop, arr):
    coeff = get_classification_values(arr)
    if prop >= coeff["max"]:
        return 1
    elif prop <= coeff["min"]:
        return 3
    else:
        return 2


# endregion


# region Student Data Process
# region Save students to Excel file
def save_students():
    _wb = openpyxl.Workbook()
    sheet = _wb["Sheet"]

    row = 2
    sheet["A1"] = "User ID"
    sheet["B1"] = "IM"
    sheet["C1"] = "CM"
    sheet["D1"] = "MM"
    sheet["E1"] = "PM"
    sheet["F1"] = "Интеллект"
    sheet["G1"] = "Креативность"
    sheet["H1"] = "Мотивация"
    sheet["I1"] = "Личность"
    for data in students:
        sheet.cell(row, 1).value = data["user_id"]
        sheet.cell(row, 2).value = data["IM"]
        sheet.cell(row, 3).value = data["CM"]
        sheet.cell(row, 4).value = data["MM"]
        sheet.cell(row, 5).value = data["PM"]
        sheet.cell(row, 6).value = data["R_IM"]
        sheet.cell(row, 7).value = data["R_CM"]
        sheet.cell(row, 8).value = data["R_MM"]
        sheet.cell(row, 9).value = data["R_PM"]

        row += 1
    _wb.save(student_result_filename)


# endregion

# region Load students data from file to array
def load_students():
    timer = timeit.default_timer()

    print("\nОбработка пользователей")

    wb = openpyxl.load_workbook(filename=student_filename, data_only=True)
    sheet = wb["CommonData"]
    print(f"Время открытия файла : {round(timeit.default_timer() - timer, 2)} сек.")

    _students = []
    IM, CM, MM, PM = [], [], [], []

    print("Student loading...")
    for i in range(2, sheet.max_row + 1):
        sys.stdout.write(f"\r{i - 1} of {sheet.max_row - 1}")

        cell = sheet.cell

        _students.append(cell(i, 1).value)
        IM.append(cell(i, 2).value)
        CM.append(cell(i, 3).value)
        MM.append(cell(i, 4).value)
        PM.append(cell(i, 5).value)
    print()

    # region Calculate values
    length = len(_students)
    IM = list(map(lambda x: round(x, 2), (rankdata(IM) / length * 0.8)))
    CM = list(map(lambda x: round(x, 2), (rankdata(CM) / length * 0.8)))
    MM = list(map(lambda x: round(x, 2), (rankdata(MM) / length * 0.8)))
    PM = list(map(lambda x: round(x, 2), (rankdata(PM) / length * 0.8)))

    print("Calculate values")
    for i in range(length):
        sys.stdout.write(f"\r{i + 1} of {length}")

        students.append({
            "user_id": _students[i],
            "communities": [],
            "IM": IM[i],
            "CM": CM[i],
            "MM": MM[i],
            "PM": PM[i],
            "R_IM": classification(IM[i], IM),
            "R_CM": classification(CM[i], CM),
            "R_MM": classification(MM[i], MM),
            "R_PM": classification(PM[i], PM)
        })
    print()
    # endregion

    save_students()

    print("Обработано записей :", length)
    print(f"Время обработки : {round((timeit.default_timer() - timer), 2)} сек.")


# endregion
# endregion

# region Community Data Process
# region Save communities to Excel file (worksheet: community_classification)
def save_communities():
    _wb = openpyxl.Workbook()
    _wb.create_sheet("community_classification")
    _wb.create_sheet("total_results")
    sheet = _wb["Sheet"]

    row = 2
    sheet["A1"] = "Group ID"
    sheet["B1"] = "I1"
    sheet["C1"] = "I2"
    sheet["D1"] = "I3"
    sheet["E1"] = "C1"
    sheet["F1"] = "C2"
    sheet["G1"] = "C3"
    sheet["H1"] = "M1"
    sheet["I1"] = "M2"
    sheet["J1"] = "M3"
    sheet["K1"] = "P1"
    sheet["L1"] = "P2"
    sheet["M1"] = "P3"

    print("Community saving...")
    for data in communities:
        sheet.cell(row, 1).value = data["group_id"]

        sheet.cell(row, 2).value = data["I1"]
        sheet.cell(row, 3).value = data["I2"]
        sheet.cell(row, 4).value = data["I3"]

        sheet.cell(row, 5).value = data["C1"]
        sheet.cell(row, 6).value = data["C2"]
        sheet.cell(row, 7).value = data["C3"]

        sheet.cell(row, 8).value = data["M1"]
        sheet.cell(row, 9).value = data["M2"]
        sheet.cell(row, 10).value = data["M3"]

        sheet.cell(row, 11).value = data["P1"]
        sheet.cell(row, 12).value = data["P2"]
        sheet.cell(row, 13).value = data["P3"]

        row += 1
    _wb.save(communities_result_filename)


# endregion

# region Load communities data from file to array
def add_communities_to_array(a):
    comms = []
    max_row = len(a)
    for i in range(max_row):
        items = a[i][~np.isnan(a[i])]
        sys.stdout.write(f"\r{i + 1} of {max_row}")

        _c = {
            "group_id": np.uint(items[0]),
            "students": []
        }
        for item in items[1:]: _c["students"].append(np.uint(item))
        comms.append(_c)
    print()
    return comms


def load_communities():
    timer = timeit.default_timer()

    print("\nОбработка подписок")

    wb = openpyxl.load_workbook(filename=communities_filename, data_only=True)
    sheet = wb["Groups"]

    pd_ws = pd.read_excel(communities_filename)
    np_array = pd_ws.to_numpy()
    print(f"Время открытия файла : {round(timeit.default_timer() - timer, 2)} сек.")

    timer = timeit.default_timer()
    print("Community loading...")
    _communities = add_communities_to_array(np_array)

    print("Community processing...")
    i = 0
    for community in _communities:
        sys.stdout.write(f"\r{i + 1} of {len(_communities)}")

        _c = {
            "group_id": community["group_id"],
            "I1": 0, "I2": 0, "I3": 0,
            "C1": 0, "C2": 0, "C3": 0,
            "M1": 0, "M2": 0, "M3": 0,
            "P1": 0, "P2": 0, "P3": 0
        }

        for student in students:
            if student["user_id"] in community["students"]:
                student["communities"].append(community["group_id"])

                if student["R_IM"] == 1:
                    _c["I1"] += 1
                elif student["R_IM"] == 2:
                    _c["I2"] += 1
                elif student["R_IM"] == 3:
                    _c["I3"] += 1

                if student["R_CM"] == 1:
                    _c["C1"] += 1
                elif student["R_CM"] == 2:
                    _c["C2"] += 1
                elif student["R_CM"] == 3:
                    _c["C3"] += 1

                if student["R_MM"] == 1:
                    _c["M1"] += 1
                elif student["R_MM"] == 2:
                    _c["M2"] += 1
                elif student["R_MM"] == 3:
                    _c["M3"] += 1

                if student["R_PM"] == 1:
                    _c["P1"] += 1
                elif student["R_PM"] == 2:
                    _c["P2"] += 1
                elif student["R_PM"] == 3:
                    _c["P3"] += 1
        communities.append(_c)
        i += 1
    print()

    save_communities()

    print("Обработано записей :", len(communities))
    print(f"Время обработки : {round((timeit.default_timer() - timer), 2)} сек.")


# endregion
# endregion

# region Community Classification Block
# region Save community classification to Excel file (worksheet: total_results)
def save_community_classifications(groups, wb):
    sheet = wb["total_results"]

    sheet["A1"] = "User ID"
    sheet["B1"] = "IM"
    sheet["C1"] = "Классификация"
    sheet["D1"] = "Кол-во маркерных сообществ"
    sheet["E1"] = "Общее кол-во сообществ"

    row = 2
    for data in groups:
        cell = sheet.cell

        cell(row, 1).value = data["user_id"]
        cell(row, 2).value = data["IM"]
        cell(row, 3).value = data["rank"]
        cell(row, 4).value = data["marker_groups"]
        cell(row, 5).value = data["total_groups"]

        row += 1


# endregion

# region Community Classification
def community_classification():
    print("\nКлассификация подписок")

    wb = openpyxl.load_workbook(communities_result_filename)
    sheet = wb["community_classification"]
    cell = sheet.cell

    sheet["A1"] = "Group ID"
    sheet["B1"] = "IM"
    sheet["C1"] = "CM"
    sheet["D1"] = "MM"
    sheet["E1"] = "PM"

    _classification = []

    row = 2
    # region Calculate values
    i = 0
    print("Group value calculating...")
    for group in communities:
        sys.stdout.write(f"\r{i + 1} of {len(communities)}")

        cell(row, 1).value = group["group_id"]

        pI1 = group["I1"] / (group["I1"] + group["I2"] + group["I3"])
        pI3 = group["I3"] / (group["I1"] + group["I2"] + group["I3"])
        piM = round(pI1 - pI3, 2)
        cell(row, 2).value = piM

        pC1 = group["C1"] / (group["C1"] + group["C2"] + group["C3"])
        pC3 = group["C3"] / (group["C1"] + group["C2"] + group["C3"])
        pcM = round(pC1 - pC3, 2)
        cell(row, 3).value = pcM

        pM1 = group["M1"] / (group["M1"] + group["M2"] + group["M3"])
        pM3 = group["M3"] / (group["M1"] + group["M2"] + group["M3"])
        pmM = round(pM1 - pM3, 2)
        cell(row, 4).value = pmM

        pP1 = group["P1"] / (group["P1"] + group["P2"] + group["P3"])
        pP3 = group["P3"] / (group["P1"] + group["P2"] + group["P3"])
        ppM = round(pP1 - pP3, 2)
        cell(row, 5).value = ppM

        _classification.append({
            "group_id": group["group_id"],
            "IM": piM,
            "CM": pcM,
            "MM": pmM,
            "PM": ppM
        })

        row += 1
    print()
    # endregion

    # region Marking Student Groups and calculate rank
    marker_groups = []
    i = 0
    print("Student groups marking...")
    for student in students:
        sys.stdout.write(f"\r{i + 1} of {len(students)}")

        _groups = student["communities"]

        marker_group = {
            "user_id": student["user_id"],
            "IM": 0,
            "rank": 0,
            "marker_groups": 0,
            "total_groups": 0
        }

        for group_id in _groups:
            _g = next((p for p in _classification if p["group_id"] == group_id), None)
            if _g is None: continue
            if _g["IM"] > 0: marker_group["marker_groups"] += 1

            marker_group["total_groups"] += 1
            marker_group["IM"] += _g["IM"]

        marker_group["IM"] /= marker_group["marker_groups"] if marker_group["marker_groups"] != 0 else 1

        marker_groups.append(marker_group)
    print()

    for group in marker_groups: group["rank"] = classification(group["IM"], [p["IM"] for p in marker_groups])
    # endregion

    print("Communities classification saving...")
    save_community_classifications(marker_groups, wb)

    wb.save(communities_result_filename)


# endregion
# endregion


# region Parsing VK User communities
def parse_vk_user_groups(access_token):
    timer = timeit.default_timer()
    # api_url = f"https://api.vk.com/method/users.getSubscriptions?access_token={access_token}&v=5.131&extended=1%count=95"
    api_url = f"https://api.vk.com/method/groups.get?access_token={access_token}&v=5.131&extended=1"
    users = []
    wb = openpyxl.load_workbook(filename=student_filename, data_only=True)
    sheet = wb["CommonData"]

    _wb = openpyxl.Workbook()
    w_sheet = _wb["Sheet"]
    w_sheet["A1"] = "User ID"
    w_sheet["B1"] = "Group ID"
    w_sheet["C1"] = "Group Name"

    max_row = sheet.max_row + 1
    try:
        for i in range(2, max_row):
            time.sleep(.20)

            print(i - 1, "of", f"{max_row}")
            user_id = sheet.cell(i, 1).value
            url = f"{api_url}&user_id={user_id}"

            request = requests.get(url)
            r = request.json()

            user = {
                "user_id": user_id,
                "groups": []
            }

            if "response" not in r:
                if "error" in r:
                    error = r["error"]
                    print(error["error_code"], error["error_msg"])
                    continue
                else:
                    print("Error")
                    break

            for r_items in r["response"]["items"]:
                if "name" not in r_items: continue

                user["groups"].append({
                    "id": r_items["id"],
                    "name": r_items["name"]
                })
            users.append(user)
    except:
        pass

    print("\nPARSING COMPLETE")

    _wb = openpyxl.Workbook()
    sheet = _wb["Sheet"]
    sheet["A1"] = "User ID"
    sheet["B1"] = "Group ID"
    sheet["C1"] = "Group Name"

    row = 2
    for user in users:
        sheet.cell(row, 1).value = user["user_id"]

        g_row = row
        for group in user["groups"]:
            sheet.cell(g_row, 2).value = group["id"]
            sheet.cell(g_row, 3).value = group["name"]
            g_row += 1

        row = g_row
    _wb.save("Data\\Test.xlsx")
    print(f"Total execution time : {round(timeit.default_timer() - timer, 2)} сек.")


# endregion


# region Excel UserGroupsModel Processing
def communities_processing():
    wb = openpyxl.load_workbook(filename="Data\\UserGroupsModel.xlsx", data_only=True)
    sheet = wb["Sheet"]
    test_sheet = wb.create_sheet("test_sheet")
    groups = []
    all_groups = []
    print("DATA PROCESSING")
    i = 2
    max_row = sheet.max_row
    while sheet.cell(i, 1).value is not None:
        print(i, "of", max_row)

        j = i + 1

        group = {
            "user_id": sheet.cell(i, 1).value,
            "groups": []
        }

        group["groups"].append(sheet.cell(i, 2).value)
        all_groups.append(sheet.cell(j, 2).value)

        while sheet.cell(j, 1).value is None:
            group["groups"].append(sheet.cell(j, 2).value)
            all_groups.append(sheet.cell(j, 2).value)
            j += 1

        i = j

        groups.append(group)
        if i == max_row: break
        # if i > 144: break
    print("\nDATA SAVING")
    i = 1
    for group_id in set(all_groups):
        print(i, "of", len(set(all_groups)))

        test_sheet.cell(i, 1).value = group_id

        j = 2
        for g in groups:
            if group_id in g["groups"]:
                test_sheet.cell(i, j).value = g["user_id"]
                j += 1

        i += 1
    del wb["Sheet"]
    wb.save("Data\\Test-2.xlsx")


# endregion


# region Program start
if __name__ == "__main__":
    print("Начало обработки :", datetime.datetime.now())
    timer = timeit.default_timer()

    load_students()
    load_communities()
    community_classification()

    # parse_vk_user_groups("your_user_token")
    # communities_processing()

    print("\n\nКонец обработки :", datetime.datetime.now())
    print(f"Общее время обработки : {round(timeit.default_timer() - timer, 2)} сек.")
# endregion
